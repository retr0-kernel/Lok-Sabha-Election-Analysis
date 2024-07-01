import requests
from bs4 import BeautifulSoup
import pandas as pd
import matplotlib.pyplot as plt
import re
import os
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

urls = [
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-S27.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-S10.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-S11.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-U09.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-U06.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-S12.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-S06.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-S07.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-S08.htm",
    "https://results.eci.gov.in/PcResultGenJune2024/partywiseresult-U08.htm"
]

def extract_state_name(url):
    return url.split('-')[-1].split('.')[0]

def plot_to_base64(plt):
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=300, bbox_inches='tight')
    buf.seek(0)
    return base64.b64encode(buf.getvalue()).decode('utf-8')

def create_dashboard(state_name, df_seats, df_vote_share, directory):
    seat_plot = df_seats.plot(kind="bar", x="Party", y="Seats", title=f"Party-wise Seat Distribution ({state_name})", legend=False, figsize=(10, 6))
    plt.ylabel("Seats Won")
    seat_plot_img = plot_to_base64(plt)
    plt.close()

    vote_share_plot = df_vote_share.plot(kind="pie", y="Votes", labels=df_vote_share["Party"], autopct='%1.1f%%', title=f"Party-wise Vote Share ({state_name})", figsize=(10, 6))
    plt.ylabel("")
    vote_share_plot_img = plot_to_base64(plt)
    plt.close()

    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Election Results Dashboard - {state_name}</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 1200px;
                margin: 0 auto;
                padding: 20px;
                background-color: #f4f4f4;
            }}
            h1, h2 {{
                color: #2c3e50;
            }}
            .dashboard-container {{
                display: flex;
                flex-wrap: wrap;
                justify-content: space-between;
            }}
            .dashboard-item {{
                flex-basis: calc(50% - 20px);
                background-color: #fff;
                border-radius: 8px;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                margin-bottom: 20px;
                padding: 20px;
                box-sizing: border-box;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                margin-bottom: 20px;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 12px;
                text-align: left;
            }}
            th {{
                background-color: #f2f2f2;
                font-weight: bold;
            }}
            img {{
                max-width: 100%;
                height: auto;
                margin-bottom: 20px;
            }}
            .toggle-btn {{
                background-color: #3498db;
                color: #fff;
                border: none;
                padding: 10px 15px;
                cursor: pointer;
                border-radius: 4px;
                font-size: 16px;
            }}
            .toggle-btn:hover {{
                background-color: #2980b9;
            }}
            .hidden {{
                display: none;
            }}
        </style>
    </head>
    <body>
        <h1>Election Results Dashboard - {state_name}</h1>
        <div class="dashboard-container">
            <div class="dashboard-item">
                <h2>Party-wise Seat Distribution</h2>
                <button class="toggle-btn" onclick="toggleVisibility('seats-table', 'seats-plot')">Toggle Table/Plot</button>
                <div id="seats-table">
                    <table>
                        <tr><th>Party</th><th>Seats</th></tr>
                        {df_seats.to_html(index=False, header=False)}
                    </table>
                </div>
                <div id="seats-plot" class="hidden">
                    <img src="data:image/png;base64,{seat_plot_img}" alt="Seat Distribution Plot">
                </div>
            </div>
            <div class="dashboard-item">
                <h2>Party-wise Vote Share</h2>
                <button class="toggle-btn" onclick="toggleVisibility('vote-table', 'vote-plot')">Toggle Table/Plot</button>
                <div id="vote-table">
                    <table>
                        <tr><th>Party</th><th>Votes</th></tr>
                        {df_vote_share.to_html(index=False, header=False)}
                    </table>
                </div>
                <div id="vote-plot" class="hidden">
                    <img src="data:image/png;base64,{vote_share_plot_img}" alt="Vote Share Plot">
                </div>
            </div>
        </div>
        <script>
            function toggleVisibility(tableId, plotId) {{
                var table = document.getElementById(tableId);
                var plot = document.getElementById(plotId);
                if (table.classList.contains('hidden')) {{
                    table.classList.remove('hidden');
                    plot.classList.add('hidden');
                }} else {{
                    table.classList.add('hidden');
                    plot.classList.remove('hidden');
                }}
            }}
        </script>
    </body>
    </html>
    """

    dashboard_path = os.path.join(directory, f'{state_name}_dashboard.html')
    with open(dashboard_path, 'w') as file:
        file.write(html_content)
    
    print(f"Dashboard created for {state_name}: {dashboard_path}")

def process_url(url):
    state_name = extract_state_name(url)  
    directory = f'./{state_name}'
    if not os.path.exists(directory):
        os.makedirs(directory)   
    response = requests.get(url)
    soup = BeautifulSoup(response.content, "html.parser")
    party_seats = []
    grid_boxes = soup.find_all("div", class_="grid-box")
    for box in grid_boxes:
        party_name = box.find("h4").text
        seats_won = box.find("h2").text
        party_seats.append({"Party": party_name, "Seats": int(seats_won)})
    
    script = soup.find("script", string=lambda text: text and "var xValues" in text).string

    x_values_pattern = re.compile(r"var xValues = \[(.*?)\];", re.DOTALL)
    y_values_pattern = re.compile(r"var yValues = \[(.*?)\];", re.DOTALL)

    x_values_match = x_values_pattern.search(script)
    y_values_match = y_values_pattern.search(script)

    x_values = x_values_match.group(1).split(',')
    y_values = y_values_match.group(1).split(',')

    x_values = [x.strip().strip("'").strip('"') for x in x_values]
    y_values = [y.strip() for y in y_values]

    x_values = [x for x in x_values if x]
    y_values = [int(y) for y in y_values if y.isdigit()]

    party_vote_share = []
    for i in range(len(x_values)):
        party_vote_share.append({"Party": x_values[i], "Votes": y_values[i]})
    
    df_seats = pd.DataFrame(party_seats)
    df_vote_share = pd.DataFrame(party_vote_share)

    seat_plot_path = os.path.join(directory, f'{state_name}_party_seat_distribution.png')
    vote_share_plot_path = os.path.join(directory, f'{state_name}_party_vote_share.png')
    
    df_seats.plot(kind="bar", x="Party", y="Seats", title=f"Party-wise Seat Distribution ({state_name})", legend=False, figsize=(10, 6))
    plt.ylabel("Seats Won")
    plt.savefig(seat_plot_path)
    plt.close()
    
    df_vote_share.plot(kind="pie", y="Votes", labels=df_vote_share["Party"], autopct='%1.1f%%', title=f"Party-wise Vote Share ({state_name})", figsize=(10, 6))
    plt.ylabel("")
    plt.savefig(vote_share_plot_path)
    plt.close()

    create_dashboard(state_name, df_seats, df_vote_share, directory)

    excel_file_path = os.path.join(directory, f'{state_name}_election_results.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "Election Results"

    ws.cell(row=1, column=1, value=f"Party-wise Seat Distribution ({state_name})").font = Font(bold=True)
    for col, header in enumerate(["Party", "Seats"], start=1):
        ws.cell(row=2, column=col, value=header).font = Font(bold=True)
    for row, data in enumerate(df_seats.itertuples(index=False), start=3):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    start_row = len(df_seats) + 5
    ws.cell(row=start_row, column=1, value=f"Party-wise Vote Share ({state_name})").font = Font(bold=True)
    for col, header in enumerate(["Party", "Votes"], start=1):
        ws.cell(row=start_row+1, column=col, value=header).font = Font(bold=True)
    for row, data in enumerate(df_vote_share.itertuples(index=False), start=start_row+2):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    img_seat = XLImage(seat_plot_path)
    img_vote = XLImage(vote_share_plot_path)
    
    ws.add_image(img_seat, 'E2')
    ws.add_image(img_vote, f'E{start_row}')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(excel_file_path)
    
    print(f"Data, plots, and Excel file saved for {state_name}")

for url in urls:
    process_url(url)